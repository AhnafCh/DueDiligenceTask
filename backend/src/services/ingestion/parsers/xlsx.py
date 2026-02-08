"""XLSX document parser using openpyxl."""
import logging
from pathlib import Path
from typing import List
from openpyxl import load_workbook

from .base import BaseParser, ParsedContent

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """Parser for Microsoft Excel spreadsheets."""
    
    def supports(self, file_extension: str) -> bool:
        """Check if file extension is .xlsx"""
        return file_extension.lower() == ".xlsx"
    
    def parse(self, file_path: Path) -> ParsedContent:
        """
        Parse XLSX document and extract data from all sheets.
        
        Args:
            file_path: Path to XLSX file
            
        Returns:
            ParsedContent with extracted text and metadata
            
        Raises:
            ValueError: If XLSX is invalid
            IOError: If file cannot be read
        """
        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
            
            sheet_texts: List[str] = []
            total_rows = 0
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Add sheet header
                sheet_content = [f"=== Sheet: {sheet_name} ==="]
                
                # Extract data from rows
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    
                    # Skip completely empty rows
                    if any(val.strip() for val in row_values):
                        row_text = " | ".join(row_values)
                        sheet_content.append(row_text)
                        total_rows += 1
                
                # Add sheet content if not empty
                if len(sheet_content) > 1:
                    sheet_texts.append("\n".join(sheet_content))
            
            # Combine all sheets
            full_text = "\n\n".join(sheet_texts)
            cleaned_text = self.clean_text(full_text)
            
            # Treat each sheet as a "page"
            page_count = len(workbook.sheetnames)
            page_texts = sheet_texts
            
            # Extract metadata
            metadata = self.extract_metadata(
                page_count=page_count,
                sheet_count=len(workbook.sheetnames),
                total_rows=total_rows,
                file_size=file_path.stat().st_size,
                file_name=file_path.name
            )
            
            # Add workbook properties if available
            try:
                props = workbook.properties
                if props.title:
                    metadata['title'] = props.title
                if props.creator:
                    metadata['creator'] = props.creator
                if props.created:
                    metadata['created'] = str(props.created)
                if props.modified:
                    metadata['modified'] = str(props.modified)
            except Exception as e:
                logger.debug(f"Could not extract XLSX metadata: {e}")
            
            workbook.close()
            
            return ParsedContent(
                text=cleaned_text,
                page_count=page_count,
                metadata=metadata,
                page_texts=page_texts,
                bounding_boxes=None  # XLSX doesn't have bounding boxes
            )
            
        except Exception as e:
            logger.error(f"Error parsing XLSX {file_path}: {e}")
            raise IOError(f"Failed to parse XLSX: {e}")
